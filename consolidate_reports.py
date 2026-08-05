import os
import glob
import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def consolidate_reports():
    unified_dir = os.path.join(os.getcwd(), "Unified_Reports")
    excel_dir = os.path.join(unified_dir, "Excel")
    screenshots_dir = os.path.join(unified_dir, "screenshots")
    logs_dir = os.path.join(unified_dir, "logs")

    for d in [unified_dir, excel_dir, screenshots_dir, logs_dir]:
        os.makedirs(d, exist_ok=True)

    modules_config = [
        {"name": "Android Mobile E2E", "dir": "automation", "report_pattern": "automation/Test Results/*.xlsx"},
        {"name": "Web Application E2E", "dir": "selenium_automation", "report_pattern": "selenium_automation/Test Results/*.xlsx"},
        {"name": "Backend Vulnerability Scan", "dir": "security_automation", "report_pattern": "security_automation/Test Results/*.xlsx"},
        {"name": "Performance Load Test", "dir": "load_automation", "report_pattern": "load_automation/Test Results/*.xlsx"},
    ]

    all_test_cases = []
    module_summaries = []

    total_executed = 0
    total_passed = 0
    total_failed = 0
    total_skipped = 0

    for mod in modules_config:
        matching_files = glob.glob(mod["report_pattern"])
        mod_passed = 0
        mod_failed = 0
        mod_skipped = 0
        mod_total = 0

        for file_path in matching_files:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                # Check for Executed Test Cases sheet
                sheet = wb["Executed Test Cases"] if "Executed Test Cases" in wb.sheetnames else wb.active
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or not any(row):
                        continue
                    test_id = str(row[0]) if len(row) > 0 and row[0] is not None else f"TC-{len(all_test_cases)+1:04d}"
                    class_name = str(row[1]) if len(row) > 1 and row[1] is not None else mod["name"]
                    test_name = str(row[2]) if len(row) > 2 and row[2] is not None else "Test Case"
                    status = str(row[3]).upper() if len(row) > 3 and row[3] is not None else "UNKNOWN"
                    duration = str(row[4]) if len(row) > 4 and row[4] is not None else "0"
                    error_msg = str(row[5]) if len(row) > 5 and row[5] is not None else "N/A"

                    if "PASS" in status:
                        mod_passed += 1
                        total_passed += 1
                        norm_status = "PASSED"
                    elif "FAIL" in status:
                        mod_failed += 1
                        total_failed += 1
                        norm_status = "FAILED"
                    else:
                        mod_skipped += 1
                        total_skipped += 1
                        norm_status = "SKIPPED"

                    mod_total += 1
                    total_executed += 1

                    all_test_cases.append({
                        "module": mod["name"],
                        "test_id": test_id,
                        "class": class_name,
                        "method": test_name,
                        "status": norm_status,
                        "duration": duration,
                        "error": error_msg,
                        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })
            except Exception as e:
                print(f"Error parsing {file_path}: {e}")

        pass_rate = (mod_passed / mod_total * 100) if mod_total > 0 else 0.0
        module_summaries.append({
            "module": mod["name"],
            "total": mod_total,
            "passed": mod_passed,
            "failed": mod_failed,
            "skipped": mod_skipped,
            "pass_rate": f"{pass_rate:.1f}%",
            "status": "PASS" if mod_failed == 0 and mod_total > 0 else ("FAIL" if mod_failed > 0 else "NO TESTS")
        })

    overall_pass_rate = (total_passed / total_executed * 100) if total_executed > 0 else 0.0

    # 1. GENERATE UNIFIED EXCEL REPORT
    master_wb = openpyxl.Workbook()

    # Sheet 1: Executive Summary
    summary_ws = master_wb.active
    summary_ws.title = "Executive Summary"
    summary_ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    summary_headers = ["Testing Module", "Total Executed", "Passed", "Failed", "Skipped", "Pass Rate", "Status"]
    summary_ws.append(summary_headers)

    for col_idx, col_name in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for mod_sum in module_summaries:
        summary_ws.append([
            mod_sum["module"],
            mod_sum["total"],
            mod_sum["passed"],
            mod_sum["failed"],
            mod_sum["skipped"],
            mod_sum["pass_rate"],
            mod_sum["status"]
        ])

    # Sheet 2: All Executed Test Cases
    details_ws = master_wb.create_sheet(title="All Executed Tests")
    details_ws.views.sheetView[0].showGridLines = True

    detail_headers = ["Module", "Test ID", "Class", "Method/Test Name", "Status", "Execution Time (ms)", "Error Message", "Timestamp"]
    details_ws.append(detail_headers)

    for col_idx, col_name in enumerate(detail_headers, 1):
        cell = details_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for tc in all_test_cases:
        details_ws.append([
            tc["module"],
            tc["test_id"],
            tc["class"],
            tc["method"],
            tc["status"],
            tc["duration"],
            tc["error"],
            tc["timestamp"]
        ])

    for ws in [summary_ws, details_ws]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    master_excel_path = os.path.join(unified_dir, "Execution_Report.xlsx")
    master_wb.save(master_excel_path)
    
    # Also save to Excel subfolder for compatibility
    master_wb.save(os.path.join(excel_dir, "Execution_Report.xlsx"))
    print(f"Unified Excel Report created at: {master_excel_path}")

    # 2. GENERATE UNIFIED HTML DASHBOARD
    git_commit = os.environ.get("GITHUB_SHA", "Local Build")[:7]
    build_id = os.environ.get("GITHUB_RUN_ID", "N/A")
    exec_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")

    if total_executed == 0:
        execution_body_html = """
        <div class="empty-state">
            <h3>No tests executed.</h3>
            <p>No valid test result artifacts were found during this execution run.</p>
        </div>
        """
    else:
        module_rows = ""
        for m in module_summaries:
            status_badge = '<span class="badge pass">PASS</span>' if m['status'] == 'PASS' else ('<span class="badge fail">FAIL</span>' if m['status'] == 'FAIL' else '<span class="badge warn">NO TESTS</span>')
            module_rows += f"""
            <tr>
                <td><strong>{m['module']}</strong></td>
                <td>{m['total']}</td>
                <td class="text-success">{m['passed']}</td>
                <td class="text-danger">{m['failed']}</td>
                <td>{m['skipped']}</td>
                <td><strong>{m['pass_rate']}</strong></td>
                <td>{status_badge}</td>
            </tr>
            """

        execution_body_html = f"""
        <div class="metrics-grid">
            <div class="card metric-card">
                <div class="metric-val">{total_executed}</div>
                <div class="metric-lbl">Total Executed</div>
            </div>
            <div class="card metric-card border-success">
                <div class="metric-val text-success">{total_passed}</div>
                <div class="metric-lbl">Passed</div>
            </div>
            <div class="card metric-card border-danger">
                <div class="metric-val text-danger">{total_failed}</div>
                <div class="metric-lbl">Failed</div>
            </div>
            <div class="card metric-card">
                <div class="metric-val text-warning">{total_skipped}</div>
                <div class="metric-lbl">Skipped</div>
            </div>
            <div class="card metric-card">
                <div class="metric-val">{overall_pass_rate:.1f}%</div>
                <div class="metric-lbl">Pass Rate</div>
            </div>
        </div>

        <div class="card card-table">
            <h2>Module Breakdown</h2>
            <table>
                <thead>
                    <tr>
                        <th>Testing Module</th>
                        <th>Total</th>
                        <th>Passed</th>
                        <th>Failed</th>
                        <th>Skipped</th>
                        <th>Pass Rate</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
                    {module_rows}
                </tbody>
            </table>
        </div>
        """

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Enterprise Automation Dashboard</title>
    <style>
        :root {{
            --bg-color: #f4f6fb;
            --card-bg: #ffffff;
            --text-main: #0b132b;
            --primary: #001a3a;
            --success: #2e7d32;
            --danger: #c62828;
            --warning: #ed6c02;
            --border: #e0e0e0;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
            background-color: var(--bg-color);
            color: var(--text-main);
            margin: 0;
            padding: 24px;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            background: var(--primary);
            color: #ffffff;
            padding: 20px 30px;
            border-radius: 12px;
            margin-bottom: 24px;
        }}
        .header h1 {{ margin: 0; font-size: 24px; font-weight: 700; }}
        .header .meta {{ font-size: 13px; opacity: 0.8; margin-top: 4px; }}
        .btn-group {{ display: flex; gap: 12px; }}
        .btn {{
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 10px 18px;
            border-radius: 8px;
            font-weight: 600;
            font-size: 14px;
            text-decoration: none;
            cursor: pointer;
            transition: all 0.2s ease;
        }}
        .btn-excel {{ background-color: #1d6f42; color: #ffffff; }}
        .btn-pdf {{ background-color: #d32f2f; color: #ffffff; }}
        .btn:hover {{ opacity: 0.9; transform: translateY(-1px); }}
        
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .card {{
            background: var(--card-bg);
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.04);
            border: 1px solid var(--border);
        }}
        .metric-card {{ text-align: center; }}
        .metric-val {{ font-size: 32px; font-weight: 800; margin-bottom: 4px; }}
        .metric-lbl {{ font-size: 12px; font-weight: 600; color: #666; text-transform: uppercase; letter-spacing: 0.5px; }}
        
        .border-success {{ border-top: 4px solid var(--success); }}
        .border-danger {{ border-top: 4px solid var(--danger); }}
        .text-success {{ color: var(--success); }}
        .text-danger {{ color: var(--danger); }}
        .text-warning {{ color: var(--warning); }}
        
        .card-table {{ overflow-x: auto; }}
        table {{ width: 100%; border-collapse: collapse; text-align: left; }}
        th, td {{ padding: 14px 16px; border-bottom: 1px solid var(--border); font-size: 14px; }}
        th {{ background: #f8f9fa; font-weight: 700; color: var(--primary); }}
        
        .badge {{
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 700;
        }}
        .badge.pass {{ background: #e8f5e9; color: var(--success); }}
        .badge.fail {{ background: #ffebee; color: var(--danger); }}
        .badge.warn {{ background: #fff3e0; color: var(--warning); }}
        
        .empty-state {{ text-align: center; padding: 40px; background: white; border-radius: 12px; }}
        .env-footer {{ margin-top: 30px; font-size: 12px; color: #777; text-align: center; }}
    </style>
</head>
<body>
    <div class="header">
        <div>
            <h1>Unified E2E Enterprise Automation Dashboard</h1>
            <div class="meta">Execution Timestamp: {exec_time} | Commit: {git_commit} | Workflow Run: #{build_id}</div>
        </div>
        <div class="btn-group">
            <a href="./Execution_Report.xlsx" download="Execution_Report.xlsx" class="btn btn-excel">📊 Download Excel Report</a>
            <a href="./Execution_Report.pdf" download="Execution_Report.pdf" class="btn btn-pdf">📄 Download PDF Report</a>
        </div>
    </div>

    {execution_body_html}

    <div class="env-footer">
        Generated dynamically by LexisAI Enterprise QA Automation Framework Engine
    </div>
</body>
</html>
"""

    html_path = os.path.join(unified_dir, "index.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Unified HTML Dashboard created at: {html_path}")

    # 3. GENERATE UNIFIED PDF REPORT
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors

        pdf_path = os.path.join(unified_dir, "Execution_Report.pdf")
        doc = SimpleDocTemplate(pdf_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#001A3A'),
            spaceAfter=12
        )

        story.append(Paragraph("Unified E2E Automation Execution Report", title_style))
        story.append(Paragraph(f"Execution Date: {exec_time} | Commit: {git_commit}", styles['Normal']))
        story.append(Spacer(1, 14))

        # Metrics Table
        data = [["Testing Module", "Total", "Passed", "Failed", "Skipped", "Pass Rate"]]
        for m in module_summaries:
            data.append([m["module"], str(m["total"]), str(m["passed"]), str(m["failed"]), str(m["skipped"]), m["pass_rate"]])

        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#001A3A')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F4F6FB')),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#CCCCCC')),
        ]))
        story.append(t)
        doc.build(story)
        print(f"Unified PDF Report created at: {pdf_path}")
    except Exception as e:
        print(f"PDF Generation note: {e}")

if __name__ == "__main__":
    consolidate_reports()
