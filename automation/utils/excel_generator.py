import os
import sys
import subprocess
from datetime import datetime

# Auto-install openpyxl if not present
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

def format_sheet(ws):
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if '\n' in val:
                val = max(val.split('\n'), key=len)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_excel_reports(results, reports_dir):
    os.makedirs(reports_dir, exist_ok=True)
    
    # Paths
    main_report_path = os.path.join(reports_dir, "Automation_Test_Report.xlsx")
    passed_report_path = os.path.join(reports_dir, "Passed_Test_Cases.xlsx")
    failed_report_path = os.path.join(reports_dir, "Failed_Test_Cases.xlsx")
    summary_report_path = os.path.join(reports_dir, "Summary_Report.xlsx")
    
    # Styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Segoe UI", size=10)
    bold_body_font = Font(name="Segoe UI", size=10, bold=True)
    
    blue_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    green_fill = PatternFill(start_color="D6EADF", end_color="D6EADF", fill_type="solid")
    red_fill = PatternFill(start_color="FAD2E1", end_color="FAD2E1", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Filters
    passed_tests = [t for t in results if t["status"] == "Passed"]
    failed_tests = [t for t in results if t["status"] == "Failed"]
    skipped_tests = [t for t in results if t["status"] == "Skipped"]
    
    total = len(results)
    passed = len(passed_tests)
    failed = len(failed_tests)
    skipped = len(skipped_tests)
    pass_rate = (passed / total * 100) if total > 0 else 0.0
    
    # ----------------------------------------------------
    # 1. Automation_Test_Report.xlsx
    # ----------------------------------------------------
    wb = openpyxl.Workbook()
    
    # Sheet 1: Executed Test Cases
    # Columns: Test ID, Module, Test Name, Status, Execution Time, Priority
    ws_exec = wb.active
    ws_exec.title = "Executed Test Cases"
    headers_exec = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]
    ws_exec.append(headers_exec)
    for col_idx, h in enumerate(headers_exec, 1):
        cell = ws_exec.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = align_center
        
    for t in results:
        ws_exec.append([
            t["testId"], t["module"], t["testName"], t["status"], round(t["executionTime"], 3), t["priority"]
        ])
        row_idx = ws_exec.max_row
        status_cell = ws_exec.cell(row=row_idx, column=4)
        status_cell.alignment = align_center
        if t["status"] == "Passed":
            status_cell.fill = green_fill
        elif t["status"] == "Failed":
            status_cell.fill = red_fill
            
        for c in range(1, 7):
            ws_exec.cell(row=row_idx, column=c).font = body_font
            ws_exec.cell(row=row_idx, column=c).border = border_thin
    format_sheet(ws_exec)
    
    # Sheet 2: Passed Tests
    ws_passed = wb.create_sheet(title="Passed Tests")
    ws_passed.append(headers_exec)
    for col_idx, h in enumerate(headers_exec, 1):
        cell = ws_passed.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = align_center
        
    for t in passed_tests:
        ws_passed.append([
            t["testId"], t["module"], t["testName"], t["status"], round(t["executionTime"], 3), t["priority"]
        ])
        row_idx = ws_passed.max_row
        ws_passed.cell(row=row_idx, column=4).fill = green_fill
        ws_passed.cell(row=row_idx, column=4).alignment = align_center
        for c in range(1, 7):
            ws_passed.cell(row=row_idx, column=c).font = body_font
            ws_passed.cell(row=row_idx, column=c).border = border_thin
    format_sheet(ws_passed)
    
    # Sheet 3: Failed Tests
    ws_failed = wb.create_sheet(title="Failed Tests")
    headers_failed = ["Test ID", "Module", "Test Name", "Status", "Failure Reason", "Execution Time (s)", "Priority"]
    ws_failed.append(headers_failed)
    for col_idx, h in enumerate(headers_failed, 1):
        cell = ws_failed.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = align_center
        
    for t in failed_tests:
        ws_failed.append([
            t["testId"], t["module"], t["testName"], t["status"], t["reason"], round(t["executionTime"], 3), t["priority"]
        ])
        row_idx = ws_failed.max_row
        ws_failed.cell(row=row_idx, column=4).fill = red_fill
        for c in range(1, 8):
            ws_failed.cell(row=row_idx, column=c).font = body_font
            ws_failed.cell(row=row_idx, column=c).border = border_thin
    format_sheet(ws_failed)
    
    # Sheet 4: Skipped Tests
    ws_skipped = wb.create_sheet(title="Skipped Tests")
    headers_skipped = ["Test ID", "Module", "Test Name", "Status", "Skip Reason", "Priority"]
    ws_skipped.append(headers_skipped)
    for col_idx, h in enumerate(headers_skipped, 1):
        cell = ws_skipped.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = align_center
        
    for t in skipped_tests:
        ws_skipped.append([
            t["testId"], t["module"], t["testName"], t["status"], t["reason"], t["priority"]
        ])
        row_idx = ws_skipped.max_row
        ws_skipped.cell(row=row_idx, column=4).fill = light_blue_fill
        for c in range(1, 7):
            ws_skipped.cell(row=row_idx, column=c).font = body_font
            ws_skipped.cell(row=row_idx, column=c).border = border_thin
    format_sheet(ws_skipped)
    
    # Sheet 5: Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    ws_metrics.append(["Metric Name", "Value"])
    ws_metrics.cell(row=1, column=1).font = header_font
    ws_metrics.cell(row=1, column=1).fill = blue_fill
    ws_metrics.cell(row=1, column=2).font = header_font
    ws_metrics.cell(row=1, column=2).fill = blue_fill
    
    metrics = [
        ("Total Test Cases", total),
        ("Passed Cases", passed),
        ("Failed Cases", failed),
        ("Skipped Cases", skipped),
        ("Pass Percentage", f"{pass_rate:.2f}%"),
        ("Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    for m, v in metrics:
        ws_metrics.append([m, v])
        r = ws_metrics.max_row
        ws_metrics.cell(row=r, column=1).font = bold_body_font
        ws_metrics.cell(row=r, column=2).font = body_font
        ws_metrics.cell(row=r, column=1).border = border_thin
        ws_metrics.cell(row=r, column=2).border = border_thin
    format_sheet(ws_metrics)
    
    # Sheet 6: Defect Summary
    ws_defects = wb.create_sheet(title="Defect Summary")
    headers_defects = ["Defect ID", "Associated Test ID", "Module", "Priority", "Failure Reason"]
    ws_defects.append(headers_defects)
    for col_idx, h in enumerate(headers_defects, 1):
        cell = ws_defects.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = blue_fill
        
    for idx, t in enumerate(failed_tests, 1):
        ws_defects.append([
            f"DF_{idx:03d}", t["testId"], t["module"], t["priority"], t["reason"]
        ])
        r = ws_defects.max_row
        ws_defects.cell(row=r, column=5).fill = red_fill
        for c in range(1, 6):
            ws_defects.cell(row=r, column=c).font = body_font
            ws_defects.cell(row=r, column=c).border = border_thin
    format_sheet(ws_defects)
    
    wb.save(main_report_path)
    
    # ----------------------------------------------------
    # 2. Passed_Test_Cases.xlsx
    # ----------------------------------------------------
    wb_passed = openpyxl.Workbook()
    ws_p = wb_passed.active
    ws_p.title = "Passed Tests"
    ws_p.append(headers_exec)
    for col_idx, h in enumerate(headers_exec, 1):
        cell = ws_p.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = blue_fill
    for t in passed_tests:
        ws_p.append([
            t["testId"], t["module"], t["testName"], t["status"], round(t["executionTime"], 3), t["priority"]
        ])
        r = ws_p.max_row
        ws_p.cell(row=r, column=4).fill = green_fill
        for c in range(1, 7):
            ws_p.cell(row=r, column=c).font = body_font
            ws_p.cell(row=r, column=c).border = border_thin
    format_sheet(ws_p)
    wb_passed.save(passed_report_path)
    
    # ----------------------------------------------------
    # 3. Failed_Test_Cases.xlsx
    # ----------------------------------------------------
    wb_failed = openpyxl.Workbook()
    ws_f = wb_failed.active
    ws_f.title = "Failed Tests"
    ws_f.append(headers_failed)
    for col_idx, h in enumerate(headers_failed, 1):
        cell = ws_f.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = blue_fill
    for t in failed_tests:
        ws_f.append([
            t["testId"], t["module"], t["testName"], t["status"], t["reason"], round(t["executionTime"], 3), t["priority"]
        ])
        r = ws_f.max_row
        ws_f.cell(row=r, column=4).fill = red_fill
        for c in range(1, 8):
            ws_f.cell(row=r, column=c).font = body_font
            ws_f.cell(row=r, column=c).border = border_thin
    format_sheet(ws_f)
    wb_failed.save(failed_report_path)
    
    # ----------------------------------------------------
    # 4. Summary_Report.xlsx
    # ----------------------------------------------------
    wb_sum = openpyxl.Workbook()
    ws_s = wb_sum.active
    ws_s.title = "Summary Metrics"
    ws_s.append(["Metric Name", "Value"])
    ws_s.cell(row=1, column=1).font = header_font
    ws_s.cell(row=1, column=1).fill = blue_fill
    ws_s.cell(row=1, column=2).font = header_font
    ws_s.cell(row=1, column=2).fill = blue_fill
    
    for m, v in metrics:
        ws_s.append([m, v])
        r = ws_s.max_row
        ws_s.cell(row=r, column=1).font = bold_body_font
        ws_s.cell(row=r, column=2).font = body_font
        ws_s.cell(row=r, column=1).border = border_thin
        ws_s.cell(row=r, column=2).border = border_thin
    format_sheet(ws_s)
    wb_sum.save(summary_report_path)
    
    print(f"Generated all Excel reports in {reports_dir}")
